from openpyxl.styles import Side, Border, Font
import openpyxl
import re
from openpyxl.styles import Side, Border, Font
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader, Template
import pdfkit

class Report:
    def __init__(self, salary_by_year: dict, vacancies_by_year: dict, salary_by_year_for_profession: dict,
                 vacancies_by_year_for_profession: dict, salary_by_city: dict, vacancies_by_city: dict,
                 profession_name: str):
        self.__salary_by_year = salary_by_year
        self.__vacancies_by_year = vacancies_by_year
        self.__salary_by_year_for_profession = salary_by_year_for_profession
        self.__vacancies_by_year_for_profession = vacancies_by_year_for_profession
        self.__salary_by_city = salary_by_city
        self.__vacancies_by_city = vacancies_by_city
        self.__profession_name = profession_name
        self.__table1 = []
        self.__table2 = []
        self.__table3 = []

    @property
    def table1(self):
        return self.__table1

    @table1.setter
    def table1(self, table1):
        self.__table1 = table1

    @property
    def table2(self):
        return self.__table2

    @table2.setter
    def table2(self, table2):
        self.__table2 = table2

    @property
    def table3(self):
        return self.__table3

    @table3.setter
    def table3(self, table3):
        self.__table3 = table3

    def generate_excel(self):
        excel_file = openpyxl.Workbook()
        excel_file.remove(excel_file["Sheet"])
        excel_file.create_sheet("Статистика по годам")
        years = [2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022]
        title1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.__profession_name}", "Количество вакансий",
                  f"Количество вакансий - {self.__profession_name}"]
        excel_file.worksheets[0].append(title1)
        for i in years:
            if i in self.__salary_by_year.keys():
                excel_file.worksheets[0].append([i, self.__salary_by_year[i], self.__salary_by_year_for_profession[i],
                                                 self.__vacancies_by_year[i],
                                                 self.__vacancies_by_year_for_profession[i]])

        for i in range(len(title1)):  excel_file.worksheets[0].cell(1, i + 1).font = Font(bold=True)

        side = Side(border_style='thin', color="FF000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for i in range(len(self.__salary_by_year.keys()) + 1):
            for j in range(len(title1)): excel_file.worksheets[0].cell(i + 1, j + 1).border = border

        dims = {}
        for row in excel_file.worksheets[0].rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
        for col, value in dims.items(): excel_file.worksheets[0].column_dimensions[col].width = value

        excel_file.create_sheet("Статистика по городам")
        t2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        excel_file.worksheets[1].append(t2)
        c1 = list(self.__salary_by_city.keys())
        c2 = list(self.__vacancies_by_city.keys())
        for i in range(len(c1)):
            excel_file.worksheets[1].append([c1[i], self.__salary_by_city[c1[i]], "",
                                             c2[i], self.__vacancies_by_city[c2[i]]])
        for i in range(len(t2)): excel_file.worksheets[1].cell(1, i + 1).font = Font(bold=True)

        for i in range(len(c1) + 1):
            for j in range(len(t2)):
                excel_file.worksheets[1].cell(i + 1, j + 1).border = border

        for i in range(2, len(c2) + 2):
            excel_file.worksheets[1].cell(i, 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

        dims = {}
        for i in excel_file.worksheets[1].rows:
            for j in i:
                if j.value: dims[j.column_letter] = max((dims.get(j.column_letter, 0), len(str(j.value)) + 2))
        for i, j in dims.items():
            excel_file.worksheets[1].column_dimensions[i].width = j

        excel_file.save("report.xlsx")
        self.__table1 = [row for row in excel_file.worksheets[0].rows]
        for row in excel_file.worksheets[1].rows:
            flag = True
            row1 = []
            row2 = []
            for cell in row:
                if (cell.value == ""):
                    flag = False
                    continue
                if (flag):
                    row1.append(cell)
                else:
                    row2.append(cell)
            self.__table2.append(row1)
            self.__table3.append(row2)

    def generate_image(self):
        figure, ax = plt.subplots(2, 2)
        width = 0.35
        # 1 график
        labels = list(self.__salary_by_year.keys())
        x = np.arange(len(labels))
        ax[0, 0].bar(x - width / 2, self.__salary_by_year.values(), width, label="Средняя З/П")
        ax[0, 0].bar(x + width / 2, self.__salary_by_year_for_profession.values(), width,
                     label=f"З/П {self.__profession_name}")
        ax[0, 0].set_title("Уровень зарплат по годам")
        ax[0, 0].set_xticks(x, labels)
        ax[0, 0].legend(fontsize=8)
        ax[0, 0].set_xticklabels(labels, rotation=90)
        for label in (ax[0, 0].get_xticklabels() + ax[0, 0].get_yticklabels()):
            label.set_fontsize(8)
        ax[0, 0].grid(axis="y")
        # 2 график
        labels = list(self.__vacancies_by_year.keys())
        x = np.arange(len(labels))
        ax[0, 1].bar(x - width / 2, self.__vacancies_by_year.values(), width, label="Количество вакансий")
        ax[0, 1].bar(x + width / 2, self.__vacancies_by_year_for_profession.values(), width,
                     label=f"Количество вакансий\n{self.__profession_name}")
        ax[0, 1].set_title("Количество вакансий по годам")
        ax[0, 1].set_xticks(x, labels)
        ax[0, 1].legend(loc="upper left", fontsize=8)
        ax[0, 1].set_xticklabels(labels, rotation=90)
        for label in (ax[0, 1].get_xticklabels() + ax[0, 1].get_yticklabels()):
            label.set_fontsize(8)
        ax[0, 1].grid(axis="y")
        # 3 график
        labels = []
        for city in list(reversed(self.__salary_by_city.keys())):
            labels.append("\n".join(re.split(r"[ -]", city)))
        x = np.arange(len(labels))
        ax[1, 0].barh(x - width / 2, list(reversed(self.__salary_by_city.values())), width)
        ax[1, 0].set_title("Уровень зарплат по городам")
        for label in (ax[1, 0].get_xticklabels() + ax[1, 0].get_yticklabels()):
            label.set_fontsize(8)
        ax[1, 0].set_yticks(x, labels, fontsize=6, horizontalalignment="right", verticalalignment="center")
        ax[1, 0].grid(axis="x")

        # 4 график
        ax[1, 1].set_title("Доля вакансий по городам")
        other = {'Другие': 1 - sum((list(self.__vacancies_by_city.values())))}
        other.update(self.__vacancies_by_city)
        self.__vacancies_by_city = other
        labels = list(self.__vacancies_by_city.keys())
        sizes = list(self.__vacancies_by_city.values())
        ax[1, 1].pie(sizes, labels=labels, textprops={'fontsize': 6})
        ax[1, 1].axis('scaled')

        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template("pdf_template.html")

        lines = '''{% for row in table -%}<tr>
        {% for cell in row -%}
        {% if row == table[0] -%}
            <th>{{cell.value}}</th>
        {% else -%}
            <td>{{cell.value}}</td>
        {% endif -%}
        {% endfor -%}</tr>
        {% endfor -%}'''
        tm = Template(lines)
        table1 = tm.render(table=self.__table1)

        table2 = tm.render(table=self.__table2)

        table3 = tm.render(table=self.__table3)

        pdf_template = template.render({"profession_name": self.__profession_name, "table1": table1, "table2": table2, "table3": table3})

        config = pdfkit.configuration(wkhtmltopdf=r"F:\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config, options={"enable-local-file-access": None})
