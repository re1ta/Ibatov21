import csv
from math import floor
from openpyxl.styles import Side, Border, Font
import openpyxl
import re
from openpyxl.styles import Side, Border, Font
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader, Template
import pdfkit

class InputConect:
    def __init__(self, data: list, profession_name: str):
        self.__data = data
        self.__profession_name = profession_name
        self.__vacancies_by_city = {}
        self.__vacancies_amount = 0

    @property
    def data(self):
        return self.__data

    @data.setter
    def data(self, data):
        self.__data = data

    @property
    def profession_name(self):
        return self.__profession_name

    @profession_name.setter
    def profession_name(self, profession_name):
        self.__profession_name = profession_name

    def get_vacancies_by_city(self):
        vbc = {} #vacancies_by_city
        for i in self.__vacancies_by_city.keys():
            if self.__vacancies_by_city[i] >= floor(self.__vacancies_amount / 100):
                vbc[i] = round(self.__vacancies_by_city[i] / self.__vacancies_amount, 4)
        vbc = dict(sorted(vbc.items(), key=lambda item: item[1], reverse=True))
        return dict(zip(list(vbc.keys())[:10], list(vbc.values())[:10]))

    def get_vacancies_by_year_for_profession(self):
        sbyfp = {} #salary_by_year_for_profession
        for i in self.data:
            if self.profession_name in i.name:
                if i.published_at in sbyfp.keys(): sbyfp[i.published_at] += 1
                else: sbyfp[i.published_at] = 1
        if (sbyfp == {}): sbyfp = {2022: 0}
        return dict(sorted(sbyfp.items(), key=lambda item: item[0]))

    def get_vacancies_by_year(self):
        vby = {}  # vacancies_by_year
        for i in self.data:
            if i.published_at in vby.keys(): vby[i.published_at] += 1
            else: vby[i.published_at] = 1
            if i.area_name in self.__vacancies_by_city.keys(): self.__vacancies_by_city[i.area_name] += 1
            else: self.__vacancies_by_city[i.area_name] = 1
            self.__vacancies_amount += 1
        return dict(sorted(vby.items(), key=lambda item: item[0]))

    def get_salary_by_year(self):
        sby = {}
        sam = {}  # salary_by_year, salary_amount
        for i in self.data:
            if i.published_at in sby.keys():
                sby[i.published_at] += (i.salary_from + i.salary_to) / 2 * currency_to_rub[i.salary_currency]
                sam[i.published_at] += 1
            else:
                sby[i.published_at] = (i.salary_from + i.salary_to) / 2 * currency_to_rub[i.salary_currency]
                sam[i.published_at] = 1
        for i in sby.keys(): sby[i] = int(sby[i] / sam[i])
        return dict(sorted(sby.items(), key=lambda item: item[0]))

    def get_salary_by_year_for_profession(self):
        sbyfp = {} #salary_by_year_for_profession
        sa = {} #salary_amount
        for i in self.data:
            if self.profession_name in i.name:
                if (i.published_at not in sbyfp.keys()):
                    sbyfp[i.published_at] = (i.salary_from + i.salary_to) / 2 * currency_to_rub[i.salary_currency]
                    sa[i.published_at] = 1
                else:
                    sbyfp[i.published_at] += (i.salary_from + i.salary_to) / 2 * currency_to_rub[i.salary_currency]
                    sa[i.published_at] += 1
        for year in sbyfp.keys(): sbyfp[year] = int(sbyfp[year] / sa[year])
        if sbyfp == {}: sbyfp = {2022: 0}
        return dict(sorted(sbyfp.items(), key=lambda item: item[0]))

    def get_salary_by_city(self):
        sbc = {} #salary_by_city
        sa = {} #salary_amount
        for i in self.data:
            if self.__vacancies_by_city[i.area_name] >= floor(self.__vacancies_amount / 100):
                if i.area_name in sbc.keys():
                    sbc[i.area_name] += (i.salary_from + i.salary_to) / 2 * currency_to_rub[i.salary_currency]
                    sa[i.area_name] += 1
                else:
                    sbc[i.area_name] = (i.salary_from + i.salary_to) / 2 * currency_to_rub[i.salary_currency]
                    sa[i.area_name] = 1
        for city in sbc.keys(): sbc[city] = int(sbc[city] / sa[city])
        sbc = dict(sorted(sbc.items(), key=lambda item: item[1], reverse=True))
        return dict(zip(list(sbc.keys())[:10], list(sbc.values())[:10]))


currency_to_rub = {
    "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "UZS": 0.0055,
    "KGS": 0.76, "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66
}

class DataSet:
    def __init__(self, file: str):
        self.__file = file
        self.__data = self.universal_csv_parser()
        self.__vacancies_objects = [Vacancy(item) for item in self.data]

    @property
    def file(self):
        return self.__file

    @file.setter
    def file(self, file):
        self.file = file

    @property
    def data(self):
        return self.__data

    @data.setter
    def data(self, data):
        self.__data = data

    @property
    def vacancies_objects(self):
        return self.__vacancies_objects

    @vacancies_objects.setter
    def vacancies_objects(self, vacancies_objects):
        self.__vacancies_objects = vacancies_objects

    def universal_csv_parser(self) -> list:
        with open(self.file, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            array1 = []
            array2 = []
            for i in reader:
                if array1 == []:
                    array1 = i
                    lineL = len(array1)
                else:
                    if ("" not in i and len(i) == lineL): array2.append(dict(zip(array1, i)))
        if array1 == []:
            print("Пустой файл")
            exit()
        return array2

class Report:
    def __init__(self, salary_by_year: dict, vacancies_by_year: dict, salary_by_year_for_profession: dict,
                 vacancies_by_year_for_profession: dict, salary_by_city: dict, vacancies_by_city: dict,
                 profession_name: str):
        self.__salary_by_year = salary_by_year
        self.__vacancies_by_year = vacancies_by_year
        self.__salary_by_year_for_profession = salary_by_year_for_profession
        self.__vacancies_by_year_for_profession = vacancies_by_year_for_profession
        self.__salary_by_city = salary_by_city
        self.__vacancies_by_city = vacancies_by_city
        self.__profession_name = profession_name
        self.__table1 = []
        self.__table2 = []
        self.__table3 = []

    @property
    def table1(self):
        return self.__table1

    @table1.setter
    def table1(self, table1):
        self.__table1 = table1

    @property
    def table2(self):
        return self.__table2

    @table2.setter
    def table2(self, table2):
        self.__table2 = table2

    @property
    def table3(self):
        return self.__table3

    @table3.setter
    def table3(self, table3):
        self.__table3 = table3

    def generate_excel(self):
        excel_file = openpyxl.Workbook()
        excel_file.remove(excel_file["Sheet"])
        excel_file.create_sheet("Статистика по годам")
        years = [2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022]
        title1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.__profession_name}", "Количество вакансий",
                  f"Количество вакансий - {self.__profession_name}"]
        excel_file.worksheets[0].append(title1)
        for i in years:
            if i in self.__salary_by_year.keys():
                excel_file.worksheets[0].append([i, self.__salary_by_year[i], self.__salary_by_year_for_profession[i],
                                                 self.__vacancies_by_year[i],
                                                 self.__vacancies_by_year_for_profession[i]])

        for i in range(len(title1)):  excel_file.worksheets[0].cell(1, i + 1).font = Font(bold=True)

        side = Side(border_style='thin', color="FF000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for i in range(len(self.__salary_by_year.keys()) + 1):
            for j in range(len(title1)): excel_file.worksheets[0].cell(i + 1, j + 1).border = border

        dims = {}
        for row in excel_file.worksheets[0].rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
        for col, value in dims.items(): excel_file.worksheets[0].column_dimensions[col].width = value

        excel_file.create_sheet("Статистика по городам")
        t2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        excel_file.worksheets[1].append(t2)
        c1 = list(self.__salary_by_city.keys())
        c2 = list(self.__vacancies_by_city.keys())
        for i in range(len(c1)):
            excel_file.worksheets[1].append([c1[i], self.__salary_by_city[c1[i]], "",
                                             c2[i], self.__vacancies_by_city[c2[i]]])
        for i in range(len(t2)): excel_file.worksheets[1].cell(1, i + 1).font = Font(bold=True)

        for i in range(len(c1) + 1):
            for j in range(len(t2)):
                excel_file.worksheets[1].cell(i + 1, j + 1).border = border

        for i in range(2, len(c2) + 2):
            excel_file.worksheets[1].cell(i, 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

        dims = {}
        for i in excel_file.worksheets[1].rows:
            for j in i:
                if j.value: dims[j.column_letter] = max((dims.get(j.column_letter, 0), len(str(j.value)) + 2))
        for i, j in dims.items():
            excel_file.worksheets[1].column_dimensions[i].width = j

        excel_file.save("report.xlsx")
        self.__table1 = [row for row in excel_file.worksheets[0].rows]
        for row in excel_file.worksheets[1].rows:
            flag = True
            row1 = []
            row2 = []
            for cell in row:
                if (cell.value == ""):
                    flag = False
                    continue
                if (flag):
                    row1.append(cell)
                else:
                    row2.append(cell)
            self.__table2.append(row1)
            self.__table3.append(row2)

    def generate_image(self):
        figure, ax = plt.subplots(2, 2)
        width = 0.35
        # 1 график
        labels = list(self.__salary_by_year.keys())
        x = np.arange(len(labels))
        ax[0, 0].bar(x - width / 2, self.__salary_by_year.values(), width, label="Средняя З/П")
        ax[0, 0].bar(x + width / 2, self.__salary_by_year_for_profession.values(), width,
                     label=f"З/П {self.__profession_name}")
        ax[0, 0].set_title("Уровень зарплат по годам")
        ax[0, 0].set_xticks(x, labels)
        ax[0, 0].legend(fontsize=8)
        ax[0, 0].set_xticklabels(labels, rotation=90)
        for label in (ax[0, 0].get_xticklabels() + ax[0, 0].get_yticklabels()):
            label.set_fontsize(8)
        ax[0, 0].grid(axis="y")
        # 2 график
        labels = list(self.__vacancies_by_year.keys())
        x = np.arange(len(labels))
        ax[0, 1].bar(x - width / 2, self.__vacancies_by_year.values(), width, label="Количество вакансий")
        ax[0, 1].bar(x + width / 2, self.__vacancies_by_year_for_profession.values(), width,
                     label=f"Количество вакансий\n{self.__profession_name}")
        ax[0, 1].set_title("Количество вакансий по годам")
        ax[0, 1].set_xticks(x, labels)
        ax[0, 1].legend(loc="upper left", fontsize=8)
        ax[0, 1].set_xticklabels(labels, rotation=90)
        for label in (ax[0, 1].get_xticklabels() + ax[0, 1].get_yticklabels()):
            label.set_fontsize(8)
        ax[0, 1].grid(axis="y")
        # 3 график
        labels = []
        for city in list(reversed(self.__salary_by_city.keys())):
            labels.append("\n".join(re.split(r"[ -]", city)))
        x = np.arange(len(labels))
        ax[1, 0].barh(x - width / 2, list(reversed(self.__salary_by_city.values())), width)
        ax[1, 0].set_title("Уровень зарплат по городам")
        for label in (ax[1, 0].get_xticklabels() + ax[1, 0].get_yticklabels()):
            label.set_fontsize(8)
        ax[1, 0].set_yticks(x, labels, fontsize=6, horizontalalignment="right", verticalalignment="center")
        ax[1, 0].grid(axis="x")

        # 4 график
        ax[1, 1].set_title("Доля вакансий по городам")
        other = {'Другие': 1 - sum((list(self.__vacancies_by_city.values())))}
        other.update(self.__vacancies_by_city)
        self.__vacancies_by_city = other
        labels = list(self.__vacancies_by_city.keys())
        sizes = list(self.__vacancies_by_city.values())
        ax[1, 1].pie(sizes, labels=labels, textprops={'fontsize': 6})
        ax[1, 1].axis('scaled')

        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template("pdf_template.html")

        lines = '''{% for row in table -%}<tr>
        {% for cell in row -%}
        {% if row == table[0] -%}
            <th>{{cell.value}}</th>
        {% else -%}
            <td>{{cell.value}}</td>
        {% endif -%}
        {% endfor -%}</tr>
        {% endfor -%}'''
        tm = Template(lines)
        table1 = tm.render(table=self.__table1)

        table2 = tm.render(table=self.__table2)

        table3 = tm.render(table=self.__table3)

        pdf_template = template.render({"profession_name": self.__profession_name, "table1": table1, "table2": table2, "table3": table3})

        config = pdfkit.configuration(wkhtmltopdf=r"F:\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config, options={"enable-local-file-access": None})

class Vacancy:
    def __init__(self, item: dict):
        self.__name = item["name"]
        self.__salary_from = float(item["salary_from"])
        self.__salary_to = float(item["salary_to"])
        self.__salary_currency = item["salary_currency"]
        self.__area_name = item["area_name"]
        self.__published_at = int(item["published_at"][:4])

    @property
    def name(self):
        return self.__name

    @name.setter
    def name(self, name):
        self.__name = name

    @property
    def salary_from(self):
        return self.__salary_from

    @salary_from.setter
    def salary_from(self, salary_from):
        self.__salary_from = salary_from

    @property
    def salary_to(self):
        return self.__salary_to

    @salary_to.setter
    def salary_to(self, salary_to):
        self.__salary_to = salary_to

    @property
    def salary_currency(self):
        return self.__salary_currency

    @salary_currency.setter
    def salary_currency(self, salary_currency):
        self.__salary_currency = salary_currency

    @property
    def area_name(self):
        return self.__area_name

    @area_name.setter
    def area_name(self, area_name):
        self.__area_name = area_name

    @property
    def published_at(self):
        return self.__published_at

    @published_at.setter
    def published_at(self, published_at):
        self.__published_at = published_at

def get_statistic():
    file_name = input("Введите название файла: ")
    profession_name = input("Введите название профессии: ")
    data = DataSet(file_name).vacancies_objects
    if not data:
        print("Нет данных")
    else:
        result = InputConect(data, profession_name)

        print(f"Динамика уровня зарплат по годам: {result.get_salary_by_year()}")
        print(f"Динамика количества вакансий по годам: {result.get_vacancies_by_year()}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {result.get_salary_by_year_for_profession()}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {result.get_vacancies_by_year_for_profession()}")
        print(f"Уровень зарплат по городам (в порядке убывания): {result.get_salary_by_city()}")
        print(f"Доля вакансий по городам (в порядке убывания): {result.get_vacancies_by_city()}")

        report = Report(result.get_salary_by_year(),
                        result.get_vacancies_by_year(),
                        result.get_salary_by_year_for_profession(),
                        result.get_vacancies_by_year_for_profession(),
                        result.get_salary_by_city(),
                        result.get_vacancies_by_city(),
                        profession_name)

        report.generate_excel()
        report.generate_image()
        report.generate_pdf()
